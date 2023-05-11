import os
from os import PathLike
from os.path import basename
from typing import Union

from tkinter.filedialog import askopenfilenames
from tkinter import Tk

import openpyxl as xl
import csv
import sqlite3 as sql

from openpyxl.utils import column_index_from_string

TABLE_NAME = "data"
DATACOLLECTION_SHEET_NAME = "Chemicals register"
DATACOLLECTION_PRODUCT_ID_COLUMN = "D"
DATACOLLECTION_MARKER_COLUMN = "AQ"
DATA_PRODUCER_COLUMN = "hersteller"
DATA_PRODUCER_ALL = "alle Hersteller"
CROSS_REFERENCES = [
    ("handelsname", "A"),
    ("cas_nr", "H"),
    ("konzentration_prozent", "G")
]



def load_worksheet(path_to_excel: Union[PathLike, str], worksheet_name: Union[str, int]):
    workbook: xl.Workbook = xl.load_workbook(path_to_excel)

    if isinstance(worksheet_name, int):
        worksheet = workbook.worksheets[worksheet_name]
    else:
        worksheet = workbook[worksheet_name]

    return workbook, worksheet


def get_worksheet_line_count(worksheet):
    max_row = 0
    for row_nr, row in enumerate(worksheet.iter_rows()):
        if any([cell.value is not None for cell in row]):
            max_row = row_nr
        else:
            pass

    return max_row


def excel_to_csv(worksheet, output_file: Union[PathLike, str]) -> None:
    with open(output_file, "w", newline="", encoding="utf-8") as out_stream:
        writer = csv.writer(out_stream, delimiter=";")

        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(row)


def parse_header(text: str) -> str:
    return text.lower().strip().\
        replace(".", "").\
        replace("-", "_").\
        replace(" ", "_").\
        replace("%", "prozent").\
        replace("ä", "ae").\
        replace("ö", "oe").\
        replace("ü", "ue")


def create_table_from_csv(connection: sql.Connection, csv_file: Union[PathLike, str], table_name: str) -> None:
    with open(csv_file, "r", encoding="utf-8") as in_stream:
        reader = csv.DictReader(in_stream, delimiter=";")
        column_headers = reader.fieldnames

    column_headers_parsed = [parse_header(header) for header in column_headers]
    print(column_headers_parsed)

    print(f"GENERATING DATABASE TABLE WITH {len(column_headers_parsed)} COLUMNS FROM HEADER OF {basename(csv_file)}... ", end="")

    connection.execute(f"CREATE TABLE {table_name} (id INTEGER PRIMARY KEY, {','.join([f'{column_header} TEXT' for column_header in column_headers_parsed])});")
    connection.commit()

    print("DONE")


def populate_table(connection: sql.Connection, csv_file: Union[PathLike, str], table_name: str) -> None:
    cursor: sql.Cursor = connection.cursor()

    with open(csv_file, "r", encoding="utf-8") as in_stream:
        reader = csv.DictReader(in_stream, delimiter=";")

        column_headers = reader.fieldnames
        column_headers_parsed = [parse_header(header) for header in column_headers]

        for row_index, row in enumerate(reader):
            print(f"INSERTING ROW {row_index + 1} OF FILE {basename(csv_file)}... ", end="")
            cursor.execute(f"INSERT INTO {table_name} ({','.join(column_headers_parsed)}) VALUES ({','.join(['?' for _ in row])});",
                           tuple(map(lambda x: str(x).strip(), row.values())))
            print("DONE")

    connection.commit()


def get_extension(filename: Union[PathLike, str]) -> str:
    ext = filename.split(".")[-1]

    if ext == filename:
        # file has no extension
        return ""
    return ext


def update_extension(filename: Union[PathLike, str], new_extension: str) -> str:
    filename_split = filename.split(".")
    filename_base = ".".join(filename_split[:-1])

    if len(filename_base) == "":
        # Original filename had no extension
        return filename + "." + new_extension

    return filename_base + "." + new_extension


def append_to_filename(filename: Union[PathLike, str], appendix: str) -> str:
    filename_split = filename.split(".")
    filename_base = ".".join(filename_split[:-1])

    if len(filename_base) == "":
        # Original filename had no extension
        return filename + appendix

    return filename_base + appendix + "." + filename_split[-1]


def main():
    connection: sql.Connection = sql.connect("db.sqlite")
    cursor: sql.Cursor = connection.cursor()

    window = Tk()
    filenames = askopenfilenames(title="Select data file(s)",
                                 defaultextension="*.csv",
                                 filetypes=[("Comma-Separated Files", "*.csv"),
                                            ("EXCEL Files", "*.xlsx")])
    window.destroy()

    if len(filenames) == 0:
        # Assume the database is already setup/populated
        print("No files provided. Assuming the database is already populated")
    else:
        # Reset table and create with new files as template
        try:
            connection.execute(f"DROP TABLE {TABLE_NAME};")
            connection.commit()
        except sql.OperationalError:
            # Table does not exist, so it does not have to be deleted
            pass

        # Convert EXCEL files into csv if not already csv
        filenames_tmp = []
        for filename in filenames:
            if get_extension(filename) == "xlsx":
                _, worksheet = load_worksheet(filename, 0)

                # Update filename to ".csv" extension and convert
                filename = update_extension(filename, "csv")
                excel_to_csv(worksheet, filename)
            filenames_tmp.append(filename)
        filenames = filenames_tmp

        create_table_from_csv(connection, filenames[0], TABLE_NAME)

        for filename in filenames:
            populate_table(connection, filename, TABLE_NAME)


    # Read data collection ('Datenerfassung') sheets
    window = Tk()
    filenames = askopenfilenames(title="Select data collection file(s)",
                                 filetypes=[("EXCEL Files", "*.xlsx"), ])
    window.destroy()

    cross_references = [(column_name, column_index_from_string(xlsx_column)-1) for (column_name, xlsx_column) in CROSS_REFERENCES]
    datacollection_product_id_column = column_index_from_string(DATACOLLECTION_PRODUCT_ID_COLUMN) - 1
    datacollection_marker_column = column_index_from_string(DATACOLLECTION_MARKER_COLUMN) - 1

    for filename in filenames:
        print(f"PROCESSING FILE {basename(filename)}... ")
        workbook, worksheet = load_worksheet(filename, DATACOLLECTION_SHEET_NAME)
        num_matches = 0

        worksheet_line_count = get_worksheet_line_count(worksheet) + 1

        for row_index, row in enumerate(worksheet.iter_rows(min_row=3, max_row=worksheet_line_count)):
            print(f"PROCESSING ROW {row_index + 3} OF {worksheet_line_count}... ", end="")

            if str(row[datacollection_product_id_column].value).strip() in ["", "None"]:
                cross_reference_request = ' AND '.join([f'{db_column} LIKE \"{row[xlsx_column].value}\"'
                                                        for (db_column, xlsx_column) in cross_references])
                cursor.execute(f"SELECT COUNT(id) FROM {TABLE_NAME} "
                               f"WHERE {DATA_PRODUCER_COLUMN} LIKE '{DATA_PRODUCER_ALL}' AND "
                               f"{cross_reference_request}")
                res = cursor.fetchone()
                if res[0] != 0:
                    num_matches += 1
                    row[datacollection_marker_column].value = "x"

                    print("MATCH FOUND")
                    continue
            print("NO MATCH")

        print(f"FINISHED PROCESSING FILE {basename(filename)}")
        if num_matches != 0:
            out_filename = append_to_filename(filename, "_x")

            num_appendix = 2
            while os.path.exists(out_filename):
                out_filename = append_to_filename(filename, "_x" + str(num_appendix))
                num_appendix += 1

            workbook.save(out_filename)

            print(f"FOUND {num_matches} MATCHES")
            print(f"WROTE CHANGES TO {out_filename}")


if __name__ == '__main__':
    main()

from tkinter.filedialog import askopenfilename, askopenfilenames
from openpyxl import load_workbook
import io
from os.path import exists
import csv
import sqlite3 as sql

# GLOBAL VARIABLES
DB_PATH = "scripts/locations.db"
TABLE_NAME = "ref"
LOCATION_COLUMNS = [22, 23, 24, 25]

from typing import List, Tuple, Dict


def create_db():
    if exists(DB_PATH):
        return sql.connect(DB_PATH)

    connection = sql.connect(DB_PATH)
    cursor = connection.cursor()

    cursor.execute(f"CREATE TABLE IF NOT EXISTS {TABLE_NAME} "
                   f"(Liegenschaft TEXT, Haus TEXT, Etage TEXT, Raum TEXT);")

    return connection


def read_reference(db_con: sql.Connection):
    filenames = askopenfilenames(title="Select reference file", filetypes=[("Comma Separated Files", "*.csv")])

    for filename in filenames:
        with open(filename, "r") as in_stream:
            ref_data = csv.DictReader(in_stream, delimiter=";")

            to_db = [(row["Liegenschaft"], row["Haus"], row["Etage"], row["RaumBez"]) for row in ref_data]

        cur = db_con.cursor()
        cur.executemany(f"INSERT INTO {TABLE_NAME} (Liegenschaft, Haus, Etage, Raum) VALUES (?, ?, ?, ?);", to_db)
    db_con.commit()


def read_locations_excel():
    filename = askopenfilename(title="Select excel file")

    with open(filename, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    wb = load_workbook(in_mem_file, data_only=True)
    ws = wb['Chemicals register']

    data = []
    for row_nr in range(3, ws.max_row+1):
        row_data = tuple([ws.cell(row=row_nr, column=cell_nr).value for cell_nr in LOCATION_COLUMNS])

        if all(map(lambda x: x is None, row_data)):
            continue

        data.append(tuple(map(str, row_data)))

    return data


def validate_rooms(connection: sql.Connection, data: List[Tuple[str]]):
    cursor = connection.cursor()
    issues: Dict[Tuple[str], List[int]] = {}

    for index, entry_data in enumerate(data):
        cursor.execute(f"SELECT * FROM {TABLE_NAME} WHERE "
                       f"Liegenschaft=? AND Haus=? AND Etage=? AND Raum=?;", entry_data)
        res = cursor.fetchone()
        if res is None:
            if entry_data in issues.keys():
                issues[entry_data].append(index + 3)
            else:
                issues[entry_data] = [index + 3]

    return issues


def print_issues(issues: Dict[Tuple[str], List[int]]):
    for (liegenschaft, haus, etage, raum), lines in issues.items():
        print(f'Ungültiger Lagerort: Liegenschaft "{liegenschaft}" / Haus "{haus}" / Etage "{etage}" / Raum "{raum}" '
              f'\n\t\tIn Zeile(n): {", ".join(map(str, lines))}')


if __name__ == '__main__':
    db = create_db()
    read_reference(db)
    locations_data = read_locations_excel()
    invalid_locations = validate_rooms(db, locations_data)
    print_issues(invalid_locations)

    db.close()

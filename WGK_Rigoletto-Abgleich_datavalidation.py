# -*- coding: utf-8 -*-
"""
Created on Fri May 13 08:48:18 2022

@author: rbienert
"""

from tkinter.filedialog import askopenfilename
from tkinter import Tk
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation  #
from openpyxl.utils import get_column_letter
import io
# import time
import datetime
import os

from pprint import pprint


def load_validations(target_ws, ws, ws_name):
    val_dict = {}
    max_row = ws.max_row

    for column_index, column in enumerate(ws.iter_cols()):
        col = [ws.cell(row=row + 1, column=column_index + 1).value for row in range(max_row)]

        col_str = get_column_letter(column_index + 1)
        header = col[0]
        col = col[1:]

        # print(col, end="\n\n")
        col = list(filter(lambda x: x not in ["", None], col))
        # print(col, end="\n\n")
        col = list(map(str, col))
        # print(col, end="\n\n")
        # print(len(col))
        # col = list(set(col))
        # print(col, end="\n\n")
        # print(len(col))
        if len(col) < 1:
            continue

        #        print(f"{header}: {col}")
        # print("-----------------------------------------------------------------------------------------------\n\n\n")

        # validation = DataValidation(type="list", formula1=f"='{','.join(col)}'", allow_blank=True, showDropDown=True)
        validation = DataValidation(type="list", formula1=f"='{ws_name}'!${col_str}$2:${col_str}${len(col) + 2}",
                                    allow_blank=True, showDropDown=False)
        target_ws.add_data_validation(validation)

        #        print(validation.formula1)

        val_dict[header.strip()] = validation

    #    pprint(val_dict)
    return val_dict


def add_validation(ws, val_dict):
    max_row = ws.max_row

    for index, column in enumerate(ws.iter_cols()):
        header = ws.cell(row=2, column=index + 1).value
        column_str = get_column_letter(index + 1)

        try:
            validation = val_dict[header.strip()]
            validation_range = column_str + str(3) + ":" + column_str + str(max_row)

            validation.add(validation_range)

        # print(f"Added {validation_range} to {header}")
        except KeyError:
            # print(f"Assignment failed for {header}")
            continue
        except AttributeError:
            # print(f"Assignment failed for {header}")
            continue


# Dateinamen abfragen
tk = Tk()
data = askopenfilename(title="Zu prüfende Daten öffnen...")
tk.destroy()

# Daten einlesen
with open(data, "rb") as f:
    in_mem_file = io.BytesIO(f.read())

wb = load_workbook(in_mem_file, data_only=True)
ws = wb['Chemicals register']

ws_selection_name = wb.sheetnames[1]
ws_selection = wb["Auswahllisten-nichts löschen!"]

now = datetime.datetime.now()
print("aktuelle Zeit: ", (now.strftime('%Y-%m-%d %H:%M:%S')))

print("\n" + "geprüfte Datei:" + "\n" + data + "\n")
# print(ws.max_row, ws.max_column)

# Finde Index von letzter nicht-leeren Zeile
max_col = ws.max_column
for row_nr, row in enumerate(ws.iter_rows()):
    if any([cell.value is not None for cell in row]):
        max_row = row_nr
        # print(row_nr, [(col_nr, cell.value, type(cell.value)) for col_nr, cell in enumerate(row) if cell.value is not None])
    else:
        pass

# Header
# print("max_row = " + str(max_row+1))
print("verwendete Spalten:")
Spalten = [1, 2, 8, 14]
for i in Spalten:
    cell_obj = ws.cell(row=2, column=i)
    print("Spalte " + str(i) + ": " + str(cell_obj.value))
print()
print(
    "\n" + "Gleicht die eingetragene Wassergefährdungsklasse mit der Rigoletto-Datenbank (Stand 12.05.2022) ab." + "\n")
print("Höhere Einstufungen werden beibehalten")
print("Der Abgleich erfolgt nur über die CAS-Nummern.")

wgk_level_dict = {
    "k.A.": -3,
    "0": -2,
    "nwg": -1,
    "awg": 0,
    "1": 1,
    "2": 2,
    "3": 3
}


def get_level(text):
    try:
        return wgk_level_dict[text]
    except KeyError:
        return -4


# # Rigoletto einlesen
# rigoletto = r"C:\Users\rbienert\Documents\Arbeitsschutz\GSM\WGK_Rigoletto-Abgleich\Rigoletto_220512.xlsx"
# with open(rigoletto, "rb") as rigo:
#     in_mem_rigo = io.BytesIO(rigo.read())
#
# wbr = load_workbook(in_mem_rigo, read_only=True, data_only=True)
# wsr = wbr['Export_Tabelle']
# Anzahl_der_Rigolettoeinträge = wsr.max_row
# # print("\n" + "verwendete Rigoletto-Spalten:")
# # Spaltenr = [5,6]
# # for i in Spaltenr:
# #     cell_objr = wsr.cell(row = 2, column = i)
# #     print("Spalte " + str(i) + ": " + str(cell_objr.value))
#
# wgk_dict = {}
# wgk_dict_duplicates = []
#
# for row in wsr.iter_rows(min_row=3, values_only=True):  # , values_only=True
#     casr = str(row[4])
#     wgkr = str(row[5])
#
#     if casr in wgk_dict.keys():
#         if get_level(wgk_dict[casr]) != get_level(wgkr):
#             wgk_dict_duplicates.append(casr)
#         continue
#
#     wgk_dict[casr] = wgkr
#
# # Entferne alle mehrfach vorkommende elemente aus dem wgk_dict dictionary
# for duplicate_cas in list(set(wgk_dict_duplicates)):
#     wgk_dict.pop(duplicate_cas)
# print("Anzahl der Rigoletto-Einträge: " + str(Anzahl_der_Rigolettoeinträge))
# print(
#     "Anzahl der gelöschten Duplikate aus der Rigoletto-Datenbank, die unterschiedliche WGK-Einstufungen bei gleicher CAS-Nummer aufweisen: " + str(
#         len(list(set(wgk_dict_duplicates)))))
# print("Anzahl der genutzten CAS-Einträge der Rigoletto-Datenbank: " + str(len(wgk_dict)) + "\n")


# Dateinamen für Referenzdaten abfragen
tk = Tk()
data_ref = askopenfilename(title="Referenzdaten öffnen...")
tk.destroy()

# Referenzdaten einlesen
with open(data_ref, "rb") as f:
    in_mem_file = io.BytesIO(f.read())

wb_ref = load_workbook(in_mem_file, data_only=True)
ws_ref = wb_ref['Chemicals register']

# Find max row in reference data
max_row = ws_ref.max_row
for row_nr, row in enumerate(ws_ref.iter_rows()):
    if any([cell.value is not None for cell in row]):
        max_row = row_nr
    else:
        pass

wgk_dict = {}
wgk_dict_duplicates = []

for row in ws_ref.iter_rows(min_row=3, max_row=max_row, values_only=True):  # , values_only=True
    casr = str(row[7])
    wgkr = str(row[13])

    if casr in wgk_dict.keys():
        if get_level(wgk_dict[casr]) != get_level(wgkr):
            wgk_dict_duplicates.append(casr)
        continue

    wgk_dict[casr] = wgkr


row_counter = 3
confirm_counter = 0
keep_counter = 0
change_counter = 0
for row_index, row in enumerate(
        ws.iter_rows(min_row=3, max_row=max_row + 1, values_only=True)):  # , values_only=True,  max_row = max_row+1
    cas = str(row[7])
    wgk = str(row[13])
    #    print(str(row_counter) + " "+cas + " " + wgk)
    row_counter = row_counter + 1

    try:
        if get_level(wgk) > -3:
            continue

        wgkr = str(wgk_dict[cas])

        try:
            if get_level(wgk) > get_level(wgkr):
                new_wgk = wgk
                # print("wgk wgkr new_wkg  " + str(wgk_level_dict[wgk]) + " " + str(wgk_level_dict[wgkr]) + " " + str(new_wgk))
            else:
                new_wgk = wgkr
        except KeyError:
            new_wgk = wgkr
        # except Exception as error:
        #     print(error)

        if get_level(wgk) == get_level(wgkr):
            #        if wgk == wgkr:
            print(str(row_counter - 1) + " " + cas + " " + wgk + " : " + wgkr + " -> " + new_wgk)
            confirm_counter += 1
        #        elif wgk > wgkr:
        elif get_level(wgk) > get_level(wgkr):
            print(str(row_counter - 1) + " " + cas + " " + wgk + " : " + wgkr + " -> " + new_wgk + " keep")
            keep_counter += 1
        else:
            print(str(row_counter - 1) + " " + cas + " " + wgk + " : " + wgkr + " -> " + new_wgk + " ==geändert==")
            change_counter += 1
        # row[13] = new_wgk
        ws.cell(row=row_index + 3, column=14).value = new_wgk
    except KeyError:
        # print(str(row_counter - 1) + " " + "not found " + cas + " " + wgk)
        # print()
        pass
    continue

    for row in wsr.iter_rows(min_row=3, max_row=10, values_only=True):  # , values_only=True
        casr = str(row[4])
        wgkr = str(row[5])
        # print(casr + wgkr)
        if cas == casr:
            print(cas + " " + wgk + " " + casr + " " + wgkr)
            row[13] = wgkr
            break
    else:
        print("not found " + cas + " " + wgk)
    # if "keine Angabe" in wgk or "k.A." in wgk or wgk == "None":
    #     print(wgk)

add_validation(ws, load_validations(ws, ws_selection, ws_selection_name))

path = os.path.dirname(data)
# print("Pfad: " + path)
fullname = os.path.basename(data)
name, file_extension = os.path.splitext(fullname)
# print(name)
# print(file_extension)
# fullname_split = fullname.split(".")
# print(fullname_split[0] + fullname_split[1])
new_data = path + "/" + name + "_WGK.xlsx"
print("\nAnzahl der Bestätigungen: " + str(confirm_counter))
print("Anzahl der Belassungen obwohl Rigoletto eine geringere Einstufung hat: " + str(keep_counter))
print("Anzahl der Änderungen: " + str(change_counter))
print("Ausgabedatei: \n" + new_data)
wb.save(filename=new_data)

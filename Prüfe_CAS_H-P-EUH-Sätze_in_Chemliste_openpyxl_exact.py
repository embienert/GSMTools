"""
Created on Wed Apr 27 09:55:19 2022
Script created by Ralf Bienert.
Pyton version used: 3.9

This script is intended to check validity of the CAS numbers and the proper formatting
 of the H-, P-, and EUH-Statements in a EXCEL-based chemical list (Gefahrstoffliste).
 
The input format of the EXCEL-sheet is the following:
    name of sheet: Chemicals register
    1. row: annotations that are skipped
    2. row: column names
    1. column: Bezeichnung/ Handelsname
    2. column: Synonyme u. weitere Bezeichnungen
    8. column: CAS-Nr.
    9. column: H-Sätze
    10. column: P-Sätze
    11. column: EUH-Sätze
"""

from tkinter.filedialog import askopenfilename
from tkinter import Tk
from openpyxl import load_workbook
import re

# Definition der erlaubten CAS-Nummern, H-, P- und EUH-Sätze
h_pattern = re.compile(r"(?:(?<!EU)H[2-4][0-9]{2}[dDfF]{0,2}(?![0-9]))"
                           r"(?:\s*\+\s*(?<!EU)H[2-4][0-9]{2}[dDfF]{0,2}(?![0-9]))*")
p_pattern = re.compile(r"(?:P[1-5][0-9]{2}(?![0-9]))(?:\s*\+\s*P[1-5][0-9]{2}(?![0-9]))*")
euh_pattern = re.compile(r"(?:EUH[0-9]{3}[dDfF]?(?![0-9]))(?:\s*\+\s*EUH[0-9]{3}[dDfF]?(?![0-9]))*")
cas_pattern = re.compile(r"[0-9]{1,8}-[0-9]{2}-[0-9](?![0-9])")

#try:
# Import
tk = Tk()
data = askopenfilename()
tk.destroy()
wb = load_workbook(filename=data, read_only=True)
ws = wb['Chemicals register']

print("\n" + "geprüfte Datei:" + "\n" + data + "\n")

print("verwendete Spalten:")
max_col = ws.max_column
max_row = ws.max_row
Spalten = [1,2,8,9,10,11]
for i in Spalten:
    cell_obj = ws.cell(row = 2, column = i)
    print("Spalte " + str(i) + ": " + cell_obj.value)
print()


# CAS
cas_error_counter = 0
row_counter = 3
for row in ws.iter_rows(min_row=3, max_row = max_row, values_only=True):
    name = row[0]
    name = name if name else ""
    cas = row[7]
    row_counter = row_counter + 1
#    cas = str(casa)
    if cas is None:
        cas = ""
    cas = (str(cas)).strip()
    if cas.strip() == "":
        continue
    if cas.strip() == "keine Angabe":
        continue
    if cas.strip() == "k.A.":
        continue
    elif re.fullmatch(cas_pattern, cas) == None:
        cas_error_counter = cas_error_counter + 1
        print("Prüfe das Format (inkl. Leerzeichen) der CAS-Nummer " + cas +" in Zeile " + str(row_counter-1) + "\t" + name)
        continue 

    cas_split = cas.split("-")
    numbers, checksum = "".join(cas_split[:-1]), int(cas_split[-1])
    cas_sum = 0
    
    for value, number in enumerate(reversed(numbers)):
        cas_sum += (value + 1) * int(number)
    
    if cas_sum % 10 != checksum:
        cas_error_counter = cas_error_counter + 1
        print("Falsche CAS-Nummer " + cas +" in Zeile " + str(row_counter-1) + "\t" + name)
   
if cas_error_counter == 0:
    print("Das Format der CAS-Nummern ist in Ordnung!")
    

# # H-Sätze
# h_error_counter = 0
# row_counter = 3
# for row in ws.iter_rows(min_row=3, max_row = max_row, values_only=True):
#     name = row[0]
#     h = row[8]
#     row_counter = row_counter + 1
    
#     if h is None:
#         h = ""
#     if str(h).strip() == "":
#         continue
    
#     if "keine Angabe" in str(h) or "k.A." in str(h):
#         continue
#     h = str(h)
#     h_split = h.split(",")
#     for i in range(0, len(h_split)):
#         h_split_i_strip = h_split[i].strip()

#         if re.match(h_pattern, h_split_i_strip) == None:
#             h_error_counter = h_error_counter + 1
#             print("Prüfe das Format (inkl. unsichtbare Zeichen) oder Gültigkeit des H-Satzes " + h_split_i_strip +" in Zeile " + str(row_counter-1) + "\t\t" + name)

# if h_error_counter == 0:
#     print("Das Format der H-Sätze ist in Ordnung!")
    
    
# # P-Sätze
# p_error_counter = 0
# row_counter = 3
# for row in ws.iter_rows(min_row=3, max_row = max_row, values_only=True):
#     name = row[0]
#     name = name if name else ""
#     p = str(row[9])
#     row_counter = row_counter + 1
    
#     if p is None:
#         p = ""
#     if p.strip() == "":
#         continue
    
#     if "keine Angabe" in p or "k.A." in p:
#         continue
    
#     p_split = p.split(",")
#     for i in range(0, len(p_split)):
#         p_split_i_strip = p_split[i].strip()

#         if re.match(p_pattern, p_split_i_strip) == None:
#             p_error_counter = p_error_counter + 1
#             try:
#                 print("Prüfe das Format (inkl. unsichtbare Zeichen) oder Gültigkeit des P-Satzes " + p_split_i_strip +" in Zeile " + str(row_counter-1) + "\t\t" + name)
#             except TypeError:
#                 print("Error")
#                 print(p_split_i_strip, type(p_split_i_strip))
#                 print(name, type(name))
    
# if p_error_counter == 0:
#     print("Das Format der P-Sätze ist in Ordnung!")
    
    
# # EUH-Sätze
# euh_error_counter = 0
# row_counter = 3
# for row in ws.iter_rows(min_row=3, max_row = max_row, values_only=True):
#     name = row[0]
#     euh = row[10]
#     row_counter = row_counter + 1
    
#     if euh is None:
#         euh = ""
#     if euh.strip() == "":
#         continue
    
#     if "keine Angabe" in euh or "k.A." in euh:
#         continue
    
#     euh_split = euh.split(",")
#     for i in range(0, len(euh_split)):
#         euh_split_i_strip = euh_split[i].strip()

#         if re.match(euh_pattern, euh_split_i_strip) == None:
#             euh_error_counter = euh_error_counter + 1
#             print("Prüfe das Format (inkl. unsichbare Zeichen) des EUH-Satzes " + euh_split_i_strip +" in Zeile " + str(row_counter-1) + "\t\t" + name)

# if euh_error_counter == 0:
#     print()
#     print("Das Format der EUH-Sätze ist in Ordnung!")


wb.close()


# except Exception as error:
#     print("Unhandled Exception:", error)


print()
input("Press ENTER to close.")



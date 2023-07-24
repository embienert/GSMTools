from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl as xl
from typing import Optional, Tuple, Dict, List, Any
import io


class Column:
    def __init__(self, title, column_name, column_header):
        self.title = title
        self.name = column_name
        self.index = column_index_from_string(self.name)
        self.header = column_header

    def __str__(self):
        return f"{self.__class__.__name__}(title={self.title}, name={self.name}, " \
               f"index={self.index}, header={self.header})"

    def __repr__(self):
        return self.__str__()


class ReferenceColumn(Column):
    def __init__(self, title, column_name, column_header, allow_empty=False, match_partial=False):
        super().__init__(title, column_name, column_header)

        self.csv_take_last = self.title == "einsatz"

        self.allow_empty = allow_empty
        self.match_partial = match_partial
        self.data: List = []

    def set_data(self, data):
        self.data = map(lambda elem: str(elem).strip(), set(data))
        self.data = list(filter(lambda x: x not in ["", str(None)], self.data))

        if self.allow_empty and str(None) not in self.data:
            self.data.append(str(None))

    def validate(self, value):
        if self.csv_take_last:
            value = str(value).split(",")[-1]
        value_cleaned = str(value).strip()

        if self.match_partial:
            for entry in self.data:
                if entry in value_cleaned:
                    return True
            return False

        return value_cleaned in self.data  # True if there is matching value in data list, False otherwise

    def get_validator(self):
        values = self.data
        if str(None) in values:
            values.remove(str(None))

        return DataValidation(type="list", formula1=";".join(map(str, values)),
                              allow_blank=self.allow_empty, showDropDown=False)


class DataColumn(Column):
    def __init__(self, title, column_name, column_header):
        super().__init__(title, column_name, column_header)

    def add_validator(self, validator: DataValidation, max_row):
        validation_range = self.name + str(3) + ":" + self.name + str(max_row)
        validator.add(validation_range)


class _Excel:
    columns: Optional[Column] = None
    sheet_name: Optional[str] = None

    def __init__(self, filename):
        self.workbook: Optional[xl.Workbook] = None
        self.worksheet: Optional[Worksheet] = None

        self._load_file(filename)

    def _load_file(self, filename):
        with open(filename, "rb") as in_stream:
            file_mem = io.BytesIO(in_stream.read())

        self.workbook = xl.load_workbook(file_mem, data_only=True)
        self.worksheet = self.workbook[self.sheet_name]

    def max_row(self):
        row_count = 0
        for row_nr, row in enumerate(self.worksheet.iter_rows()):
            if any([cell.value is not None for cell in row]):
                row_count = row_nr

        return row_count

    def __str__(self):
        return f"{self.__class__.__name__}(workbook_path={self.workbook.path}, worksheet={self.sheet_name})"


class Reference(_Excel):
    columns: List[ReferenceColumn] = [
        ReferenceColumn("reinheit", "A", "Reinheit", allow_empty=True),
        ReferenceColumn("cmr_kategorie", "B", "bei CMR-Stoffen \nKategorie eintragen", allow_empty=True),
        ReferenceColumn("wgk", "C", "WGK", allow_empty=True),
        ReferenceColumn("abwasser", "D", "gelangt das Produkt in das Abwasser "),
        ReferenceColumn("liegenschaft", "E", "Liegenschaft"),
        ReferenceColumn("haus", "F", "Gebäude/ Haus"),
        ReferenceColumn("standort", "G", "Standort im Raum"),
        ReferenceColumn("einsatz", "H", "Einsatzort und Verwendungszweck (u.a. für Kopf der BA) \nvorher: "
                                        "\"Verfahren, Anwendung\""),
        ReferenceColumn("verwendungstyp", "I", "Verwendungstypen / Produktkategorie"),
        ReferenceColumn("lagerklasse", "J", "Lagerklasse", allow_empty=True),
        ReferenceColumn("antragssteller", "K", "Antragsteller*in (Name, Vorname)"),
        ReferenceColumn("ansprechpartner", "L", "Ansprechpartner*in vor Ort = Gefahrstoffkoordinator*in (Name, Vorname)"),
        ReferenceColumn("fuehrungskraft", "M", "Führungskraft (Name, Vorname)")
    ]
    sheet_name = "Auswahllisten-nichts löschen!"
    min_row = 2

    def __init__(self, filename):
        super().__init__(filename)

        self._row_count = self.max_row()
        self._column_titles = map(lambda x: x.title, self.columns)
        self._column_headers = map(lambda x: x.header, self.columns)

        self.load_data()

    def load_data(self):
        for ref_col in self.columns:
            self.get_column_data(ref_col)

    def get_column_data(self, col: ReferenceColumn):
        data = []
        for row_nr in range(self.min_row, self._row_count+1):
            data.append(self.worksheet.cell(row=row_nr, column=col.index).value)

        col.set_data(data)

    def column_by_title(self, title):
        matches = list(filter(lambda x: x.title == title, self.columns))

        if len(matches) == 0:
            return None
        elif len(matches) > 1:
            raise LookupError("Multiple columns with same title.")

        return matches[0]

    def column_by_header(self, header):
        matches = list(filter(lambda x: x.header == header, self.columns))

        if len(matches) == 0:
            return None
        elif len(matches) > 1:
            raise LookupError("Multiple columns with same header.")

        return matches[0]

    def get_column_titles(self):
        return list(self._column_titles)

    def get_column_headers(self):
        return list(self._column_headers)

    def __getattr__(self, item):
        if item in self._column_titles:
            return self.column_by_title(item)

        return self.__dict__[item]


class Data(_Excel):
    columns: List[DataColumn] = [
        DataColumn("bezeichnung", "A", "Bezeichnung/ Handelsname"),
        DataColumn("synonyme", "B", "Synonyme u. weitere Bezeichnungen"),
        DataColumn("hersteller", "C", "Hersteller "),
        DataColumn("artikelnummer", "D", "Artikelnummer des Herstellers"),
        DataColumn("zusammensetzung", "E", "Gemisch,\nZusammen-\nsetzung"),
        DataColumn("reinheit", "F", "Reinheit"),
        DataColumn("konzentration", "G", "Konzentration\nin %"),
        DataColumn("cas", "H", "CAS-Nr."),
        DataColumn("h", "I", "H-Sätze"),
        DataColumn("p", "J", "P-Sätze"),
        DataColumn("euh", "K", "EUH-Sätze"),
        DataColumn("cmr", "L", "CMR-Marker"),
        DataColumn("cmr_kategorie", "M", "bei CMR-Stoffen \nKategorie eintragen"),
        DataColumn("wgk", "N", "WGK"),
        DataColumn("abwasser", "O", "gelangt das Produkt in das Abwasser"),
        DataColumn("dichte", "P", "Dichte in  g/cm3 (kg/L)"),
        DataColumn("inhalt_liq_gas_l", "Q", "Flüssigkeiten /  Gase Behälterinhalt NUR  in L"),
        DataColumn("inhalt_liq_kg", "R", "Flüssigkeiten Behälterinhalt umgerechnet in kg"),
        DataColumn("inhalt_fest_kg", "S", "Feststoffe Behälterinhalt NUR in kg "),
        DataColumn("anz_behaelter", "T", "Anzahl der Behälter"),
        DataColumn("beschaffungsmenge", "U", "max. Beschaffungs-menge "),
        DataColumn("liegenschaft", "V", "Liegenschaft"),
        DataColumn("haus", "W", "Gebäude/ Haus"),
        DataColumn("etage", "X", "Etage"),
        DataColumn("raum", "Y", "Raum"),
        DataColumn("standort", "Z", "Standort im Raum"),
        DataColumn("standort_zusatz", "AA", "Zusatz Standort im Raum"),
        DataColumn("einsatz", "AB", "Einsatzort und Verwendungszweck (u.a. für Kopf der BA) \nvorher: \"Verfahren, "
                                    "Anwendung\""),
        DataColumn("verwendungstyp", "AC", "Verwendungstypen / Produktkategorie"),
        DataColumn("lagerhinweise", "AD", "Lagerungshinweise"),
        DataColumn("lagerklasse", "AE", "Lagerklasse"),
        DataColumn("antragsteller", "AF", "Antragsteller*in (Name, Vorname)"),
        DataColumn("ansprechpartner", "AG", "Ansprechpartner*in vor Ort = Gefahrstoffkoordinator*in (Name, Vorname)"),
        DataColumn("fuehrungskraft", "AH", "Führungskraft (Name, Vorname)"),
        DataColumn("fb", "AI", "FB"),
        DataColumn("kommentare", "AJ", "Kommentare")
    ]
    sheet_name = "Chemicals register"
    min_row = 3

    def __init__(self, filename):
        super().__init__(filename)

        self._row_count = self.max_row() + 1
        self._column_titles = map(lambda x: x.title, self.columns)
        self._column_headers = map(lambda x: x.header, self.columns)

    def column_by_title(self, title):
        matches = list(filter(lambda x: x.title == title, self.columns))

        if len(matches) == 0:
            return None
        elif len(matches) > 1:
            raise LookupError("Multiple columns with same header.")

        return matches[0]

    def column_by_header(self, header):
        matches = list(filter(lambda x: x.header == header, self.columns))

        if len(matches) == 0:
            return None
        elif len(matches) > 1:
            raise LookupError("Multiple columns with same header.")

        return matches[0]

    def get_column_titles(self):
        return list(self._column_titles)

    def get_column_headers(self):
        return list(self._column_headers)

    def get_associated_columns(self, reference: Reference,
                               subset_header: Optional[List[str]] = None,
                               subset_title: Optional[List[str]] = None):
        data_reference_assoc = None

        if subset_title is not None:
            reference_titles = reference.get_column_titles()
            titles_intersect = list(filter(lambda x: x in reference_titles and x in subset_title, self._column_titles))
            data_reference_assoc = [(self.column_by_title(title), reference.column_by_title(title)) for title in
                                    titles_intersect]

        if data_reference_assoc is None:
            reference_headers = reference.get_column_headers()
            subset_header = subset_header if subset_header is not None else reference_headers
            headers_intersect = list(filter(lambda x: x in reference_headers and x in subset_header,
                                            self._column_headers))

            data_reference_assoc = [(self.column_by_header(header), reference.column_by_header(header)) for header in
                                    headers_intersect]

        return data_reference_assoc

    def cross_reference(self, reference: Reference,
                        subset_header: Optional[List[str]] = None,
                        subset_title: Optional[List[str]] = None):
        data_reference_assoc = self.get_associated_columns(reference, subset_header, subset_title)

        # Initialize empty dictionary for storing found discrepancies
        issues: Dict[Column, List[Tuple[int, Any]]] = {}

        for data_col, reference_col in data_reference_assoc:
            # Iterate over all pairs of associated DataColumn and ReferenceColumn objects
            column = data_col.index  # Get column index for data worksheet
            issues[data_col] = []  # Initialize found issues with empty list

            for row_nr in range(self.min_row, self._row_count+1):  # Iterate over each row
                value = self.worksheet.cell(row=row_nr, column=column).value  # Get cell value for current row and col

                if not reference_col.validate(value):
                    # No matching value found in reference -> Add to list of row numbers and values with issues
                    issues[data_col].append((row_nr, value))

        return issues

    def add_data_validations(self, reference: Reference,
                             subset_header: Optional[List[str]] = None,
                             subset_title: Optional[List[str]] = None):
        data_reference_assoc = self.get_associated_columns(reference, subset_header, subset_title)

        for data_col, reference_col in data_reference_assoc:
            validator = reference_col.get_validator()
            self.worksheet.add_data_validation(validator)
            data_col.add_validator(validator, self._row_count)

    def save_copy(self, filename: str):
        self.workbook.save(filename=filename)

    def __getattr__(self, item):
        if item in self._column_titles:
            return self.column_by_title(item)

        return self.__dict__[item]


# if __name__ == '__main__':
#     while True:
#         try:
#             print(exec(input("> ")))
#         except Exception as error:
#             print(error)

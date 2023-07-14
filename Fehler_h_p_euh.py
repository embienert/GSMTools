"""
Created on Wed May 05 09:55:19 2022
Script created by Ralf Bienert.
Pyton version used: 3.9

This script is intended to add the letter in front of the numbers of the H-, P-, and EUH-Statements
 in a EXCEL-based chemical list (Gefahrstoffliste).
 
The input format of the EXCEL-sheet is the following:
    name of sheet: Chemicals register
    1. row: annotations that are skipped
    2. row: column names
    1. column: Bezeichnung/ Handelsname
    2. column: Synonyme u. weitere Bezeichnungen
    8. column: CAS-Nr.
    9. column: H-Sätze
    10. column: P-Sätze
    11. column: EUH-Sätze
"""

from tkinter.filedialog import askopenfilename
from tkinter import Tk
from openpyxl import load_workbook
import io
import time
import datetime
#import re

# H-, P, EUH-Sätze 
ohne_h_list = ["200", "201", "202", "203", "204", "205", "206", "207", "208", "220", "221", "222", "223", "224", "225", "226", "228", "229", "230", "231", "232", "240", "241", "242", "250", "251", "252", "260", "261", "270", "271", "272", "280", "281", "290", "300", "301", "302", "304", "310", "311", "312", "314", "315", "317", "318", "319", "330", "331", "332", "334", "335", "336", "340", "341", "350", "350i", "351", "360", "360D", "360F", "360FD", "360Df", "360Fd", "361", "361d", "361f", "361fd", "362", "370", "371", "372", "373", "300 + 310", "300 + 330", "310 + 330", "300 + 310 + 330", "301 + 311", "301 + 331", "311 + 331", "301 + 311 + 331", "302 + 312", "302 + 332", "312 + 332", "302 + 312 + 332", "400", "410", "411", "412", "413", "420"]
mit_h_list = ["H200", "H201", "H202", "H203", "H204", "H205", "H206", "H207", "H208", "H220", "H221", "H222", "H223", "H224", "H225", "H226", "H228", "H229", "H230", "H231", "H232", "H240", "H241", "H242", "H250", "H251", "H252", "H260", "H261", "H270", "H271", "H272", "H280", "H281", "H290", "H300", "H301", "H302", "H304", "H310", "H311", "H312", "H314", "H315", "H317", "H318", "H319", "H330", "H331", "H332", "H334", "H335", "H336", "H340", "H341", "H350", "H350i", "H351", "H360", "H360D", "H360F", "H360FD", "H360Df", "H360Fd", "H361", "H361d", "H361f", "H361fd", "H362", "H370", "H371", "H372", "H373", "H300 + H310", "H300 + H330", "H310 + H330", "H300 + H310 + H330", "H301 + H311", "H301 + H331", "H311 + H331", "H301 + H311 + H331", "H302 + H312", "H302 + H332", "H312 + H332", "H302 + H312 + H332", "H400", "H410", "H411", "H412", "H413", "H420"]

ohne_p_list = ["101", "102", "103", "201", "202", "210", "211", "212", "220", "222", "223", "230", "231", "232", "233", "234", "235", "240", "241", "242", "243", "244", "250", "251", "260", "261", "262", "263", "264", "270", "271", "272", "273", "280", "282", "283", "284", "301", "302", "303", "304", "305", "306", "308", "310", "311", "312", "313", "314", "315", "320", "321", "330", "331", "332", "333", "334", "335", "336", "337", "338", "340", "342", "351", "352", "353", "360", "361", "362", "363", "364", "370", "371", "372", "373", "375", "376", "377", "378", "380", "381", "390", "391", "401", "402", "403", "404", "405", "406", "407", "410", "411", "412", "413", "420", "422", "501", "502", "503", "221", "281", "285", "307", "309", "322", "341", "350", "374"]
mit_p_list = ["P101", "P102", "P103", "P201", "P202", "P210", "P211", "P212", "P220", "P222", "P223", "P230", "P231", "P232", "P233", "P234", "P235", "P240", "P241", "P242", "P243", "P244", "P250", "P251", "P260", "P261", "P262", "P263", "P264", "P270", "P271", "P272", "P273", "P280", "P282", "P283", "P284", "P301", "P302", "P303", "P304", "P305", "P306", "P308", "P310", "P311", "P312", "P313", "P314", "P315", "P320", "P321", "P330", "P331", "P332", "P333", "P334", "P335", "P336", "P337", "P338", "P340", "P342", "P351", "P352", "P353", "P360", "P361", "P362", "P363", "P364", "P370", "P371", "P372", "P373", "P375", "P376", "P377", "P378", "P380", "P381", "P390", "P391", "P401", "P402", "P403", "P404", "P405", "P406", "P407", "P410", "P411", "P412", "P413", "P420", "P422", "P501", "P502", "P503", "P221", "P281", "P285", "P307", "P309", "P322", "P341", "P350", "P374"] 
# Die folgenden hier gelisteten P-Sätze sind bereits aufgehoben, werden aber noch akzeptiert: , "P221", "P242", "P281", "P285", "P307", "P309", "P322", "P341", "P350", "P374"

ohne_euh_list = ["014", "018", "019", "029", "031", "032", "044", "066", "070", "071", "201", "201A", "202", "203", "204", "205", "206", "207", "208", "209", "209A", "210", "211", "212", "401"]
mit_euh_list = ["EUH014", "EUH018", "EUH019", "EUH029", "EUH031", "EUH032", "EUH044", "EUH066", "EUH070", "EUH071", "EUH201", "EUH201A", "EUH202", "EUH203", "EUH204", "EUH205", "EUH206", "EUH207", "EUH208", "EUH209", "EUH209A", "EUH210", "EUH211", "EUH212", "EUH401", "EUH001", "EUH006", "EUH059"]


# Dateinamen abfragen
tk = Tk()
data = askopenfilename()
tk.destroy()


# Daten einlesen
with open(data, "rb") as f:
    in_mem_file = io.BytesIO(f.read())
    
wb = load_workbook(in_mem_file, read_only=True, data_only=True)
ws = wb['Chemicals register']


now = datetime.datetime.now()
print("aktuelle Zeit: ", (now.strftime('%Y-%m-%d %H:%M:%S')))

print("\n" + "geprüfte Datei:" + "\n" + data + "\n")
#print(ws.max_row, ws.max_column)


# Finde Index von letzter nicht-leeren Zeile
max_col = ws.max_column
for row_nr, row in enumerate(ws.iter_rows()):
    if any([cell.value is not None for cell in row]):
        max_row = row_nr
        # print(row_nr, [(col_nr, cell.value, type(cell.value)) for col_nr, cell in enumerate(row) if cell.value is not None])
    else:
        pass
        
    
# Header
print("max_row = " + str(max_row+1))
print("\n" + "verwendete Spalten:")
Spalten = [1,2,8,9,10,11]
for i in Spalten:
    cell_obj = ws.cell(row = 2, column = i)
    print("Spalte " + str(i) + ": " + str(cell_obj.value))
print()
print("\n" +"Prüft, ob die H-, P- und EUH-Sätze mit den jeweiligen Buchstaben beginnen und gültig sind."+"\n")


def plus_buchstabe(spalte, ohne_list, mit_list, Header):
    print(Header)
    row_counter = 3

    change_counter = 0
    for row in ws.iter_rows(min_row=3, max_row = max_row+1, values_only=True):  # , values_only=True
        h = str(row[spalte])
        row_counter = row_counter + 1
        
        if h == "None":
            h = ""
            #print(h)
            continue
        
        if "keine Angabe" in h or "k.A." in h or "-" in h:
            #print(h)
            continue
        
        h_split = h.split(",")
        for i in range(0, len(h_split)):
            result = []
            for plus_split in h_split[i].split("+"):
#                if plus_split.strip() not in mit_list and plus_split.strip() not in ohne_list:
                if plus_split.strip() not in mit_list:
                    print(f"Zeile {row_counter-1}: fehlender Buchstabe oder ungültige Ziffernfolge: " + plus_split.strip())
                    change_counter = change_counter + 1
                    continue
                
                for k in range(0, len(ohne_list)):
                    if ohne_list[k] == plus_split.strip():
                        result.append(mit_list[k])
                        change_counter = change_counter + 1
                        break
                else:
                    result.append(plus_split.strip())
                    #plus_split = mit_list[k]
                # result = list(set(result))
            h_split[i] = " + ".join(result)
                        
                    
        #print(", ".join(h_split))
 
    #print("  ENDE letzte Zeile: " + str(row_counter-1))
    print()
    return change_counter


if plus_buchstabe(8, ohne_h_list, mit_h_list, "Fehler in H-Sätzen") == 0:
    print("Alle H-Sätze beginnen bereits mit H." + "\n")
    
if plus_buchstabe(9, ohne_p_list, mit_p_list, "Fehler in P-Sätzen") == 0:
    print("Alle P-Sätze beginnen bereits mit P." + "\n")
    
if plus_buchstabe(10, ohne_euh_list, mit_euh_list, "Fehler in EUH-Sätzen") == 0:
    print("Alle EUH-Sätze beginnen bereits mit EUH."+ "\n")

wb._archive.close()

#print("\n" +"Speicher deine Gefahrstoffliste in eine separate Datei bevor du Änderungen vornimmst!" + "\n" + "Die Ausgabe dieses Skripts kann nun in die EXCEL-Datei kopiert werden." + "\n")
print("gültige H-, P- und EUH-Sätze: https://www.gefahrstoffdaten.de oder https://de.wikipedia.org/wiki/H-_und_P-S%C3%A4tze" + "\n")

time.sleep(.1)
input("Press ENTER to close. This is mandatory to start a new session.")
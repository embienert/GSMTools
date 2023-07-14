from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter import Tk

from openpyxl.utils import column_index_from_string
import openpyxl

import sqlite3 as sql

import os
import io


DATABASE_PATH = "pers.db"

REFERENCE_SHEET_NAME = "Tabelle1"
REFERENCE_ID_COLUMN = "A"
REFERENCE_NAME_COLUMN = "B"
REFERENCE_FIRSTNAME_COLUMN = "C"

GSM_SHEET_NAME = "Chemicals register"
GSM_NAME_COLUMNS = ["AF", "AG", "AH"]


def create_db(db_path):
    if os.path.exists(db_path):
        return sql.connect(db_path)

    connection = sql.connect(db_path)
    cursor = connection.cursor()

    cursor.execute(f"CREATE TABLE IF NOT EXISTS pers "
                   f"(persID INTEGER UNIQUE, name TEXT, firstname TEXT);")

    return connection


def clean_db(db: sql.Connection):
    cur = db.cursor()

    # cur.execute("SELECT name, firstname FROM pers GROUP BY name, firstname HAVING count(*) > 1;")
    # print("Cleanup: ", cur.fetchall())

    # Delete all entries with equal name and firstname
    cur.execute("DELETE FROM pers WHERE (name, firstname) IN "
                "(SELECT name, firstname FROM pers GROUP BY name, firstname HAVING count(*) > 1);")

    db.commit()


def load_reference(path, db: sql.Connection):
    data_columns = [
        column_index_from_string(REFERENCE_ID_COLUMN),
        column_index_from_string(REFERENCE_NAME_COLUMN),
        column_index_from_string(REFERENCE_FIRSTNAME_COLUMN)
    ]

    with open(path, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    wb = openpyxl.load_workbook(in_mem_file, data_only=True)
    ws = wb[REFERENCE_SHEET_NAME]

    data = []
    for row_nr in range(2, ws.max_row+1):
        row_data = tuple([ws.cell(row=row_nr, column=col_nr).value for col_nr in data_columns])

        if all(map(lambda x: x is None, row_data)):
            continue

        data.append(tuple(map(str, row_data)))

    cur = db.cursor()
    cur.executemany("INSERT OR IGNORE INTO pers (persID, name, firstname) VALUES (?, ?, ?);", data)
    db.commit()


def get_replace(db: sql.Connection, name_raw):
    try:
        name, firstname = tuple(map(lambda x: x.strip(), str(name_raw).split(",")))
    except ValueError:
        print(f"Could not resolve {name_raw} into first and last name.")
        return None

    cur = db.cursor()
    cur.execute("SELECT persID FROM pers WHERE name LIKE (?) AND firstname LIKE (?);", (name, firstname))
    results = cur.fetchall()

    if len(results) != 1:
        print(f"Could not find match for {name_raw} in reference table.")
        return None

    return results[0][0]


def main():
    data_columns = [column_index_from_string(col) for col in GSM_NAME_COLUMNS]

    # Get path to reference file from user
    root = Tk()
    reference_file = askopenfilename(title="Select reference file")
    root.destroy()

    # Get or create database
    db = create_db(DATABASE_PATH)

    # Load personnel data from reference file if path is valid
    if os.path.exists(reference_file):
        load_reference(reference_file, db)

    # Clean database from duplicate values
    clean_db(db)

    # Get path to main GSM file from user
    root = Tk()
    gsm_file = askopenfilename(title="Select GSM file")
    root.destroy()

    # Load GSM file
    with open(gsm_file, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    wb = openpyxl.load_workbook(in_mem_file, data_only=True)
    ws = wb[GSM_SHEET_NAME]

    # Process rows of GSM file
    failed = []
    for row_nr in range(3, ws.max_row+1):
        for col_nr in data_columns:
            target_cell = ws.cell(row=row_nr, column=col_nr)

            replacement = get_replace(db, target_cell.value)
            if replacement is None:
                # Less or more than one comma in the name
                failed.append(f"{target_cell.column_letter}{row_nr}")

                continue

            target_cell.value = replacement

    # Get output path from user
    root = Tk()
    out_path = asksaveasfilename(title="Save processed file")
    root.destroy()

    # print("Failed: ", failed)

    # Save processed GSM file
    wb.save(out_path)


if __name__ == '__main__':
    main()
